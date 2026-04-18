from dataclasses import dataclass
from pathlib import Path
from threading import Thread
from typing import Callable, Dict, List, Optional
import traceback

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl


@dataclass(frozen=True)
class ProcessorConfig:
    input_file: str = "sentencasNicolas.xlsx"
    output_file: str = "sentencasNicolasfeito.xlsx"
    start_row: int = 3
    max_row_scan: int = 11993
    col_hist: int = 6
    col_ses: int = 7
    col_result: int = 8


@dataclass
class HistoryData:
    ses: str = ""
    af: str = ""
    ct: str = ""
    ata: str = ""
    has_af: bool = False
    has_ct: bool = False
    has_ata: bool = False


class HistoryParser:
    @staticmethod
    def tokenize(text: str) -> List[str]:
        return text.replace("_", " ").replace(":", "").replace(".", "").lower().split()

    def parse(self, tokens: List[str], row: int) -> HistoryData:
        data = HistoryData()
        self._parse_ses(tokens, data, row)
        self._parse_af(tokens, data)
        self._parse_ct(tokens, data)
        self._parse_ata(tokens, data)
        return data

    @staticmethod
    def _parse_ses(tokens: List[str], data: HistoryData, row: int) -> None:
        if "ses" not in tokens:
            return

        ses_parts = tokens[tokens.index("ses") + 1].split("/")
        try:
            data.ses = str(int(ses_parts[0])) + "/" + ses_parts[1]
        except (IndexError, ValueError):
            print(f"\nLinha {row}: Formato de SES invalido encontrado: {ses_parts}.\n")

    @staticmethod
    def _parse_af(tokens: List[str], data: HistoryData) -> None:
        if "af" not in tokens:
            return

        for index in range(tokens.index("af"), len(tokens)):
            if tokens[index].replace("/", "").isdigit():
                data.af = tokens[index]
                data.has_af = True
                break

    @staticmethod
    def _parse_ct(tokens: List[str], data: HistoryData) -> None:
        if "ct" not in tokens:
            return

        for index in range(tokens.index("ct"), len(tokens)):
            if tokens[index].replace("/", "").isdigit():
                data.ct = tokens[index]
                data.has_ct = True
                break

    @staticmethod
    def _parse_ata(tokens: List[str], data: HistoryData) -> None:
        if "ata" not in tokens:
            return

        for index in range(tokens.index("ata"), len(tokens)):
            if tokens[index].replace("/", "").isdigit():
                data.ata = tokens[index]
                data.has_ata = True


class SpreadsheetSentenceProcessor:
    def __init__(
        self,
        config: ProcessorConfig,
        logger: Optional[Callable[[str], None]] = None,
    ) -> None:
        self.config = config
        self.parser = HistoryParser()
        self.log = logger if logger is not None else print
        self.counter_empty = 0
        self.counter_multiple_values = 0
        self.ata_to_af: Dict[str, str] = {}
        self.last_processed_row = self.config.start_row - 1

        self.workbook = openpyxl.load_workbook(self.config.input_file)
        self.worksheet = self.workbook.active

    def run(self) -> None:
        self._print_start_info()
        self._process_rows()
        self._resolve_pending_ata_values()
        self._count_and_report_column_issues()
        self._save()

    def _print_start_info(self) -> None:
        self.log("Planilha carregada com sucesso!")
        self.log(f"Planilha ativa: {self.worksheet.title}\n")

    def _process_rows(self) -> None:
        consecutive_null_history = 0
        for row in range(self.config.start_row, self.config.max_row_scan):
            self.last_processed_row = row
            if self._process_single_row(row):
                consecutive_null_history += 1
                if consecutive_null_history >= 5:
                    self.log(
                        f"Linha {row}: Encerrando processamento apos 5 linhas seguidas com historico nulo."
                    )
                    break
            else:
                consecutive_null_history = 0

        self.log("\nLeitura concluida.\n")

    def _process_single_row(self, row: int) -> bool:
        history = self.worksheet.cell(row=row, column=self.config.col_hist).value
        if self._is_null_like(history):
            self.log(f"Linha {row}: Coluna 6 (Historico) esta vazia.")
            return True

        tokens = self.parser.tokenize(str(history))
        parsed = self.parser.parse(tokens, row)
        self._apply_row_results(row, parsed)
        return False

    @staticmethod
    def _is_null_like(value: object) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        return False

    def _apply_row_results(self, row: int, parsed: HistoryData) -> None:
        if parsed.ses:
            self.worksheet.cell(row=row, column=self.config.col_ses, value=f"SES: {parsed.ses}")

        if parsed.has_af:
            self.worksheet.cell(row=row, column=self.config.col_result, value=f"AF: {parsed.af}")

        if parsed.has_ct:
            self.worksheet.cell(row=row, column=self.config.col_result, value=f"CT: {parsed.ct}")

        if parsed.has_ata and parsed.has_af and not parsed.has_ct:
            self.ata_to_af[parsed.ata] = parsed.af

        if parsed.has_ata and not parsed.has_af:
            self.worksheet.cell(row=row, column=self.config.col_result, value=f"ATA: {parsed.ata}")

        if parsed.has_af and parsed.has_ct:
            self.worksheet.cell(
                row=row,
                column=self.config.col_result,
                value=f"CT: {parsed.ct} | AF: {parsed.af}",
            )
            self.log(
                f"Linha {row}: Ambos AF ({parsed.af}) e CT ({parsed.ct}) encontrados. "
                "Favor verificar manualmente."
            )

    def _resolve_pending_ata_values(self) -> None:
        for row in self._iter_processed_rows():
            cell_value = self.worksheet.cell(row=row, column=self.config.col_result).value
            if cell_value is None or "ATA:" not in cell_value:
                continue

            ata_key = cell_value.split()[1]
            self.worksheet.cell(
                row=row,
                column=self.config.col_result,
                value=self.ata_to_af.get(ata_key, ""),
            )

    def _count_and_report_column_issues(self) -> None:
        for row in self._iter_processed_rows():
            cell_value = self.worksheet.cell(row=row, column=self.config.col_result).value
            self._report_cell_status(row, cell_value)

        self.log(str(self.ata_to_af.get("292/2025")))
        self.log(f"\nTotal de celulas sem valor na coluna 8: {self.counter_empty}")
        self.log(
            "Total de celulas com mais de um valor na coluna 8: "
            f"{self.counter_multiple_values}"
        )

    def _report_cell_status(self, row: int, cell_value: Optional[str]) -> None:
        if cell_value is None:
            self.counter_empty += 1
            return

        if len(str(cell_value).split()) > 2:
            self.counter_multiple_values += 1
            self.log(f"Linha {row}: Mais de um valor encontrado na celula: ({cell_value})")

    def _iter_processed_rows(self) -> range:
        if self.last_processed_row < self.config.start_row:
            return range(0)
        return range(self.config.start_row, self.last_processed_row + 1)

    def _save(self) -> None:
        self.workbook.save(self.config.output_file)
        self.log(f"Arquivo salvo em: {self.config.output_file}")


class SpreadsheetProcessorApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Processador de Planilhas")
        self.geometry("900x620")
        self.minsize(820, 520)
        default = ProcessorConfig()
        self.input_var = tk.StringVar(value=default.input_file)
        self.output_var = tk.StringVar(value=default.output_file)
        self.start_row_var = tk.StringVar(value=str(default.start_row))
        self.max_row_var = tk.StringVar(value=str(default.max_row_scan))

        self.run_button: Optional[ttk.Button] = None
        self.log_text: Optional[tk.Text] = None
        self._build_ui()

    def _build_ui(self) -> None:
        container = ttk.Frame(self, padding=14)
        container.pack(fill="both", expand=True)

        form = ttk.Frame(container)
        form.pack(fill="x")
        form.columnconfigure(1, weight=1)

        ttk.Label(form, text="Arquivo de entrada (.xlsx):").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(form, textvariable=self.input_var).grid(row=0, column=1, sticky="ew", pady=6)
        ttk.Button(form, text="Selecionar", command=self._select_input_file).grid(row=0, column=2, padx=(8, 0), pady=6)

        ttk.Label(form, text="Arquivo de saida (.xlsx):").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(form, textvariable=self.output_var).grid(row=1, column=1, sticky="ew", pady=6)
        ttk.Button(form, text="Salvar como", command=self._select_output_file).grid(row=1, column=2, padx=(8, 0), pady=6)

        ttk.Label(form, text="Linha inicial:").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(form, textvariable=self.start_row_var, width=14).grid(row=2, column=1, sticky="w", pady=6)

        ttk.Label(form, text="Maximo de linhas para leitura:").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(form, textvariable=self.max_row_var, width=14).grid(row=3, column=1, sticky="w", pady=6)

        self.run_button = ttk.Button(container, text="Executar processamento", command=self._start_processing)
        self.run_button.pack(anchor="w", pady=(12, 10))

        ttk.Label(container, text="Logs:").pack(anchor="w")
        log_frame = ttk.Frame(container)
        log_frame.pack(fill="both", expand=True, pady=(6, 0))

        self.log_text = tk.Text(log_frame, wrap="word", height=20)
        self.log_text.pack(side="left", fill="both", expand=True)
        self.log_text.config(state="disabled")

        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def _select_input_file(self) -> None:
        selected = filedialog.askopenfilename(
            title="Selecione a planilha de entrada",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if not selected:
            return

        self.input_var.set(selected)
        current_output = self.output_var.get().strip()
        if not current_output or current_output == ProcessorConfig().output_file:
            self.output_var.set(self._suggest_output_path(selected))

    def _select_output_file(self) -> None:
        selected = filedialog.asksaveasfilename(
            title="Escolha o arquivo de saida",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            initialfile=Path(self.output_var.get().strip() or "saida.xlsx").name,
        )
        if selected:
            self.output_var.set(selected)

    @staticmethod
    def _suggest_output_path(input_path: str) -> str:
        source = Path(input_path)
        return str(source.with_name(f"{source.stem}_feito.xlsx"))

    def _start_processing(self) -> None:
        input_file = self.input_var.get().strip()
        output_file = self.output_var.get().strip()

        if not input_file:
            messagebox.showerror("Erro", "Informe o arquivo de entrada.")
            return
        if not Path(input_file).exists():
            messagebox.showerror("Erro", "Arquivo de entrada nao encontrado.")
            return
        if not output_file:
            messagebox.showerror("Erro", "Informe o arquivo de saida.")
            return

        try:
            start_row = int(self.start_row_var.get().strip())
            max_row_scan = int(self.max_row_var.get().strip())
        except ValueError:
            messagebox.showerror("Erro", "Linha inicial e maximo de linhas devem ser numeros inteiros.")
            return

        if start_row < 1:
            messagebox.showerror("Erro", "Linha inicial deve ser maior ou igual a 1.")
            return
        if max_row_scan <= start_row:
            messagebox.showerror("Erro", "Maximo de linhas deve ser maior que a linha inicial.")
            return

        self._clear_log()
        self._append_log("Iniciando processamento...\n")
        if self.run_button is not None:
            self.run_button.config(state="disabled")

        config = ProcessorConfig(
            input_file=input_file,
            output_file=output_file,
            start_row=start_row,
            max_row_scan=max_row_scan,
        )
        worker = Thread(target=self._process_worker, args=(config,), daemon=True)
        worker.start()

    def _process_worker(self, config: ProcessorConfig) -> None:
        try:
            processor = SpreadsheetSentenceProcessor(config=config, logger=self._append_log)
            processor.run()
            self._append_log("\nProcessamento finalizado com sucesso.")
            self.after(0, lambda: messagebox.showinfo("Sucesso", "Processamento concluido."))
        except Exception as exc:
            error_message = f"Falha ao processar: {exc}"
            self._append_log("\nFalha no processamento.")
            self._append_log(f"Erro: {exc}")
            self._append_log(traceback.format_exc())
            self.after(0, lambda: messagebox.showerror("Erro", error_message))
        finally:
            self.after(0, self._enable_run_button)

    def _enable_run_button(self) -> None:
        if self.run_button is not None:
            self.run_button.config(state="normal")

    def _clear_log(self) -> None:
        if self.log_text is None:
            return

        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

    def _append_log(self, message: str) -> None:
        self.after(0, self._append_log_main_thread, message)

    def _append_log_main_thread(self, message: str) -> None:
        if self.log_text is None:
            return

        self.log_text.config(state="normal")
        self.log_text.insert("end", f"{message}\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")


def main() -> None:
    app = SpreadsheetProcessorApp()
    app.mainloop()


if __name__ == "__main__":
    main()